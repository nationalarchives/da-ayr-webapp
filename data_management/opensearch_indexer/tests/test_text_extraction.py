import tempfile
from pathlib import Path
from unittest.mock import ANY, patch

import pytest
from docx import Document
from openpyxl import Workbook
from opensearch_indexer.text_extraction import (
    TextExtractionStatus,
    add_text_content,
    extract_text,
)


class TestExtractText:
    @pytest.mark.parametrize(
        "file_name, file_type, expected_output",
        [
            (
                "multiline.txt",
                "x-fmt/111",
                "This is line 1\nThis is line 2\nThis is line 3\nThis is line 4, the final line.\n",
            ),
            (
                "multiline.docx",
                "fmt/412",
                "This is line 1\n\n\t\t\t\t\t\t\t\t\t\t\t\t\t"
                "This is line 2\n\n\t\t\t\t\t\t\t\t\t\t\t\t\t"
                "This is line 3\n\n"
                "This is line 4, the final line.",
            ),
            (
                "multiline.doc",
                "fmt/40",
                "\nExpected content\nSecond line\nThird line\n",
            ),
            (
                "multiline.pdf",
                "fmt/276",
                "This is line 1\nThis is line 2\nThis is line 3\nThis is line 4, the final line.\n\n\x0c",
            ),
            (
                "multiline.odt",
                "fmt/291",
                "This is line 1\nThis is line 2\nThis is line 3\nThis is line 4, the final odt line.\n",
            ),
            (
                "multiline.html",
                "x-fmt/394",
                "\nThis is line 1\n            This is line 2\n            This is line 3, the final html line.\n",
            ),
            (
                "multiline.htm",
                "x-fmt/394",
                "\nThis is line 1\n            This is line 2\n            This is line 3, the final htm line.\n",
            ),
            (
                "multiline.eml",
                "fmt/278",
                "This is line 1\nThis is line 2\nThis is line 3\nThis is line 4, the final eml line.\n",
            ),
            (
                "multiline.msg",
                "x-fmt/430",
                "\n\nThis is line 1\r\nThis is line 2\r\nThis is line 3\r\nThis is line 4, the final msg line.\r\n",
            ),
            (
                "multiline.csv",
                "x-fmt/18",
                "\tHeader1\tHeader2\tHeader3\t\t\t\n\tValue1\tValue2\tValue3\t\t\t",
            ),
            (
                "multiline.xlsx",
                "fmt/214",
                "\nHeader1 Header2 Header3\nValue1 Value2 Value3\n",
            ),
        ],
    )
    def test_extract_text(self, file_name, file_type, expected_output):
        path = Path(__file__).parent / f"test_files/{file_name}"
        assert extract_text(str(path), file_type, 1) == expected_output


class TestRealTextExtraction:
    def test_textract_real_docx(self, tmp_path: Path):
        docx_path = tmp_path / "test.docx"
        doc = Document()
        doc.add_paragraph("testing textract")
        doc.save(docx_path)

        file_dict = {"file_id": "123-123", "file_puid": "fmt/412"}
        file_bytes = docx_path.read_bytes()

        extracted_text = add_text_content(file_dict, file_bytes)

        assert "testing textract" in extracted_text["content"].lower()

    def test_textract_real_xlsx(self, tmp_path: Path):
        xlsx_path = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "Header1"
        ws["B1"] = "Header2"
        ws["A2"] = "Value1"
        ws["B2"] = "Value2"

        wb.save(xlsx_path)

        file_dict = {"file_id": "693-9382", "file_puid": "fmt/214"}
        file_bytes = xlsx_path.read_bytes()

        extracted_text = add_text_content(file_dict, file_bytes)

        assert "header1" in extracted_text["content"].lower()
        assert "header2" in extracted_text["content"].lower()
        assert "value1" in extracted_text["content"].lower()
        assert "value2" in extracted_text["content"].lower()


# Mock ENVIRONMENT for slack alerts
@pytest.fixture(autouse=True)
def patch_environment():
    with patch("opensearch_indexer.text_extraction.ENVIRONMENT", "test-env"):
        yield


# Mock the extract_text function to simulate text extraction behavior
@pytest.fixture
def mock_extract_text():
    with patch("opensearch_indexer.text_extraction.extract_text") as mock:
        yield mock


# Test for successfully extracting text from a supported file type
def test_add_text_content_success(mock_extract_text, caplog):
    """
    Given a supported file type and a valid file stream,
    When text extraction succeeds,
    Then the content is updated and the status is set to SUCCEEDED.
    """

    # Given
    file = {
        "file_id": 1,
        "file_name": "example.pdf",
        "file_extension": "pdf",
        "file_puid": "fmt/276",
        "content": "",
        "text_extraction_status": "",
    }
    file_stream = b"Some file content"
    mock_extract_text.return_value = "Extracted text"

    # When
    result = add_text_content(file, file_stream)

    # Then
    assert result["content"] == "Extracted text"
    assert (
        result["text_extraction_status"] == TextExtractionStatus.SUCCEEDED.value
    )
    mock_extract_text.assert_called_once_with(ANY, "fmt/276", file["file_id"])

    assert "Text extraction succeeded for file 1" in caplog.text


def test_add_text_content_no_ffid_metadata_success(mock_extract_text, caplog):
    """
    Given a supported file type and a valid file stream and without FFID metadata,
    When text extraction succeeds,
    Then the content is updated and the status is set to SUCCEEDED.
    """

    # Given
    file = {
        "file_id": 1,
        "file_name": "example.pdf",
        "file_extension": None,
        "file_puid": "fmt/276",
        "content": "",
        "text_extraction_status": "",
    }
    file_stream = b"Some file content"
    mock_extract_text.return_value = "Extracted text"

    # When
    result = add_text_content(file, file_stream)

    # Then
    assert result["content"] == "Extracted text"
    assert (
        result["text_extraction_status"] == TextExtractionStatus.SUCCEEDED.value
    )
    mock_extract_text.assert_called_once_with(ANY, "fmt/276", file["file_id"])

    assert "Text extraction succeeded for file 1" in caplog.text


# Test for unsupported file type
def test_add_text_content_unsupported_format(caplog):
    """
    Given an unsupported file type,
    When text extraction is skipped,
    Then the content is set to an empty string and the status is set to SKIPPED.
    """

    # Given
    file = {
        "file_id": 2,
        "file_name": "example.exe",  # Unsupported file type
        "file_extension": "exe",
        "file_puid": "fmt/100",
        "content": "",
        "text_extraction_status": "",
    }
    file_stream = b"Some content that won't be extracted"

    # When
    result = add_text_content(file, file_stream)

    # Then
    assert result["content"] == ""
    assert (
        result["text_extraction_status"] == TextExtractionStatus.SKIPPED.value
    )

    assert (
        "Text extraction skipped for file 2 due to unsupported file type: fmt/100"
        in caplog.text
    )


# Test for text extraction failure
def test_add_text_content_failure(mock_extract_text, caplog):
    """
    Given a supported file type and a failing text extraction,
    When text extraction fails due to an error,
    Then the content is set to an empty string and the status is set to FAILED.
    """

    # Given
    file = {
        "file_id": 3,
        "file_name": "example.txt",  # Supported file type
        "file_extension": "txt",
        "file_puid": "x-fmt/111",
        "content": "",
        "text_extraction_status": "",
    }
    file_stream = b"Some content"

    # Simulate a failure in text extraction
    mock_extract_text.side_effect = Exception("Text extraction failed")

    # When
    with caplog.at_level("ERROR"):
        result = add_text_content(file, file_stream)

    # Then
    assert result["content"] == ""
    assert result["text_extraction_status"] == TextExtractionStatus.FAILED.value
    mock_extract_text.assert_called_once_with(ANY, "x-fmt/111", file["file_id"])

    assert (
        "Text extraction failed for file 3: Text extraction failed"
        in caplog.text
    )


# Test for file type without extension
def test_add_text_content_no_extension():
    """
    Given a file without an extension,
    When trying to extract text,
    Then the file is skipped and the status is set to SKIPPED.
    """

    # Given
    file = {
        "file_id": 4,
        "file_name": "example",  # No file extension
        "file_extension": "",  # no extension
        "file_puid": "",
        "content": "",
        "text_extraction_status": "",
    }
    file_stream = b"Some content"

    # When
    result = add_text_content(file, file_stream)

    # Then
    assert result["content"] == ""
    assert (
        result["text_extraction_status"] == TextExtractionStatus.SKIPPED.value
    )


def test_add_text_content_fallback_success():
    """
    Given a supported file type that initially fails textract,
    When fallback conversion succeeds,
    Then the content is extracted from the converted file and status is SUCCEEDED.
    """
    file = {
        "file_id": 5,
        "file_name": "example.xls",  # In fallback map
        "file_extension": "xls",
        "file_puid": "fmt/59",
        "content": "",
        "text_extraction_status": "",
    }
    file_stream = b"original xls content"

    with patch(
        "opensearch_indexer.text_extraction.textract.process"
    ) as mock_textract, patch(
        "opensearch_indexer.text_extraction.convert_file_with_libreoffice"
    ) as mock_convert:
        # Simulate first textract failure, second success after conversion
        mock_textract.side_effect = [
            Exception("initial fail"),
            b"converted content",
        ]
        mock_convert.return_value = "/tmp/example.xlsx"

        result = add_text_content(file, file_stream)

        assert result["content"] == "converted content"
        assert (
            result["text_extraction_status"]
            == TextExtractionStatus.SUCCEEDED.value
        )
        assert mock_textract.call_count == 2
        mock_convert.assert_called_once()


def test_extract_text_libreoffice_conversion_failure():
    """
    Given textract fails on the original file,
    And LibreOffice conversion also fails (e.g. subprocess error),
    Then extract_text should raise the conversion exception,
    with the initial textract failure as the cause.
    """
    file_bytes = b"dummy content"

    with tempfile.NamedTemporaryFile(suffix=".xls", delete=True) as temp:
        temp.write(file_bytes)
        temp.flush()

        with patch(
            "opensearch_indexer.text_extraction.textract.process"
        ) as mock_textract, patch(
            "opensearch_indexer.text_extraction.convert_file_with_libreoffice"
        ) as mock_convert:

            mock_textract.side_effect = Exception("initial textract failed")
            mock_convert.side_effect = Exception(
                "libreoffice conversion failed"
            )

            with pytest.raises(Exception) as exc_info:
                extract_text(temp.name, "fmt/59", 1)

            assert "libreoffice conversion failed" in str(exc_info.value)
            assert exc_info.value.__cause__ is not None
            assert "initial textract failed" in str(exc_info.value.__cause__)

            mock_textract.assert_called_once()
            mock_convert.assert_called_once()


def test_extract_text_fallback_conversion_failure():
    """
    Given textract fails on both original and converted files,
    Then extract_text should raise the converted textract exception,
    with the initial textract failure as the cause.
    """
    file_bytes = b"dummy file content"
    converted_path = "/tmp/converted.xlsx"

    with tempfile.NamedTemporaryFile(suffix=".xls", delete=True) as temp:
        temp.write(file_bytes)
        temp.flush()

        with patch(
            "opensearch_indexer.text_extraction.textract.process"
        ) as mock_textract, patch(
            "opensearch_indexer.text_extraction.convert_file_with_libreoffice"
        ) as mock_convert:

            # Given
            mock_textract.side_effect = [
                Exception("initial textract failed"),
                Exception("converted textract failed"),
            ]
            mock_convert.return_value = converted_path

            # Then
            with pytest.raises(Exception) as exc_info:
                extract_text(temp.name, "fmt/59", 1)

            assert "converted textract failed" in str(exc_info.value)
            assert exc_info.value.__cause__ is not None
            assert "initial textract failed" in str(exc_info.value.__cause__)

            assert mock_textract.call_count == 2
            mock_convert.assert_called_once()


def test_add_text_content_proactive_conversion_success(caplog):
    """
    Given an unsupported file type that is in the proactive conversion map,
    When conversion and subsequent extraction succeed,
    Then the content is extracted from the converted file and status is SUCCEEDED.
    """
    # Given a PUID in the convert map (e.g., x-fmt/44: .wpd to .docx)
    file = {
        "file_id": 6,
        "file_name": "example.wpd",
        "file_extension": "wpd",
        "file_puid": "x-fmt/44",
        "content": "",
        "text_extraction_status": "",
    }
    file_stream = b"original wpd content"

    with patch(
        "opensearch_indexer.text_extraction.textract.process"
    ) as mock_textract, patch(
        "opensearch_indexer.text_extraction.convert_file_with_libreoffice"
    ) as mock_convert:
        mock_textract.return_value = b"extracted from docx"
        mock_convert.return_value = "/tmp/example.docx"

        # When
        result = add_text_content(file, file_stream)

        # Then
        assert result["content"] == "extracted from docx"
        assert (
            result["text_extraction_status"]
            == TextExtractionStatus.SUCCEEDED.value
        )
        mock_convert.assert_called_once_with(ANY, "docx")
        mock_textract.assert_called_once_with("/tmp/example.docx")
        assert "Converting 6 to docx" in caplog.text
        assert "Text extraction succeeded for file 6" in caplog.text


def test_add_text_content_proactive_conversion_failure(caplog):
    """
    Given an unsupported file type that is in the proactive conversion map,
    When conversion fails,
    Then the status is set to FAILED and content is empty.
    """
    # Given a PUID in the convert map (e.g., fmt/126: .ppt to .pptx)
    file = {
        "file_id": 7,
        "file_name": "example.ppt",
        "file_extension": "ppt",
        "file_puid": "fmt/126",
        "content": "",
        "text_extraction_status": "",
    }
    file_stream = b"original ppt content"

    with patch(
        "opensearch_indexer.text_extraction.convert_file_with_libreoffice"
    ) as mock_convert:
        mock_convert.side_effect = Exception("conversion failed")

        # When
        result = add_text_content(file, file_stream)

        # Then
        assert result["content"] == ""
        assert (
            result["text_extraction_status"]
            == TextExtractionStatus.FAILED.value
        )
        mock_convert.assert_called_once_with(ANY, "pptx")
        assert "Converting 7 to pptx" in caplog.text
        assert (
            "Text extraction failed for file 7: conversion failed"
            in caplog.text
        )


def test_add_text_content_proactive_conversion_extract_failure(caplog):
    """
    Given an unsupported file type that is in the proactive conversion map,
    When conversion succeeds but extraction on converted file fails,
    Then the status is set to FAILED and content is empty.
    """
    # Given a PUID in the convert map (e.g., x-fmt/258: .vsd to .pdf)
    file = {
        "file_id": 8,
        "file_name": "example.vsd",
        "file_extension": "vsd",
        "file_puid": "x-fmt/258",
        "content": "",
        "text_extraction_status": "",
    }
    file_stream = b"original vsd content"

    with patch(
        "opensearch_indexer.text_extraction.textract.process"
    ) as mock_textract, patch(
        "opensearch_indexer.text_extraction.convert_file_with_libreoffice"
    ) as mock_convert:
        mock_textract.side_effect = Exception("extract after convert failed")
        mock_convert.return_value = "/tmp/example.pdf"

        # When
        result = add_text_content(file, file_stream)

        # Then
        assert result["content"] == ""
        assert (
            result["text_extraction_status"]
            == TextExtractionStatus.FAILED.value
        )
        mock_convert.assert_called_once_with(ANY, "pdf")
        mock_textract.assert_called_once_with("/tmp/example.pdf")
        assert "Converting 8 to pdf" in caplog.text
        assert (
            "Text extraction failed for file 8: extract after convert failed"
            in caplog.text
        )


# def test_add_text_content_skipped_alert():
#     """
#     Given an unsupported file type,
#     When text extraction is skipped,
#     A slack alert is sent.
#     """
#     file = {
#         "file_id": "abc123",
#         "file_name": "unsupported.xyz",
#         "file_extension": "xyz",
#         "content": "",
#         "text_extraction_status": "SKIPPED",
#     }
#     file_stream = b"some content"

#     with patch(
#         "opensearch_indexer.text_extraction.send_slack_alert"
#     ) as mock_alert:

#         result = add_text_content(file, file_stream)

#         assert result["content"] == ""
#         assert (
#             result["text_extraction_status"]
#             == TextExtractionStatus.SKIPPED.value
#         )

#         mock_alert.assert_called_once_with(
#             "Text extraction *SKIPPED* for file `abc123`\n"
#             "*Environment:* `TEST-ENV`\n"
#             "*Reason:* Unsupported file type - `xyz`"
#         )


# def test_add_text_content_failed_alert(mock_extract_text):
#     """
#     Given a supported file type and a failing text extraction,
#     When text extraction fails due to an error,
#     A slack alert is sent.
#     """
#     file = {
#         "file_id": "def456",
#         "file_name": "example.pdf",
#         "file_extension": "pdf",
#         "content": "",
#         "text_extraction_status": "FAILED",
#     }
#     file_stream = b"some content"

#     mock_extract_text.side_effect = Exception("something broke")

#     with patch(
#         "opensearch_indexer.text_extraction.send_slack_alert"
#     ) as mock_alert:

#         result = add_text_content(file, file_stream)

#         assert result["content"] == ""
#         assert (
#             result["text_extraction_status"]
#             == TextExtractionStatus.FAILED.value
#         )

#         mock_alert.assert_called_once_with(
#             "Text extraction *FAILED* for file `def456`\n"
#             "*Environment:* `TEST-ENV`\n"
#             "*Reason:* `something broke`"
#         )
